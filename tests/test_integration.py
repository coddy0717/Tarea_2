"""Pruebas de integración: verifican que data_loader, processor y
report_generator trabajen correctamente en conjunto a través de la capa
`main`, sin pasar por la interfaz de línea de comandos como proceso aparte
(eso lo cubren las pruebas end-to-end en test_e2e.py).
"""

from report_automation.main import build_arg_parser, main, run

VALID_CSV = """date,category,product,quantity,unit_price
2025-01-05,Electrónica,Laptop,2,850.00
2025-01-08,Electrónica,Mouse,10,15.50
2025-01-09,Oficina,Papel,not-a-number,4.25
"""


def test_run_generates_both_reports(tmp_path):
    csv_path = tmp_path / "transactions.csv"
    csv_path.write_text(VALID_CSV, encoding="utf-8")
    output_dir = tmp_path / "output"

    result = run(csv_path, output_dir)

    assert result["excel_path"].exists()
    assert result["html_path"].exists()
    assert result["skipped_rows"] == 1
    assert result["summary"]["total_transactions"] == 2


def test_main_returns_zero_on_success(tmp_path, capsys):
    csv_path = tmp_path / "transactions.csv"
    csv_path.write_text(VALID_CSV, encoding="utf-8")
    output_dir = tmp_path / "output"

    exit_code = main(["--input", str(csv_path), "--output-dir", str(output_dir)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "Reporte generado" in captured.out


def test_main_returns_one_when_input_missing(tmp_path, capsys):
    missing_csv = tmp_path / "no_existe.csv"

    exit_code = main(["--input", str(missing_csv)])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "[ERROR]" in captured.err


def test_run_propagates_processor_output_into_excel_report(tmp_path):
    """Verifica que los KPIs calculados por `processor` lleguen intactos
    hasta el archivo Excel generado por `report_generator`, es decir, que
    la integración entre ambos módulos no pierda ni distorsione datos."""
    import openpyxl

    csv_path = tmp_path / "transactions.csv"
    csv_path.write_text(VALID_CSV, encoding="utf-8")
    output_dir = tmp_path / "output"

    result = run(csv_path, output_dir)

    workbook = openpyxl.load_workbook(result["excel_path"])
    category_sheet = workbook["Categorías"]
    assert category_sheet.cell(row=2, column=1).value == "Electrónica"
    assert category_sheet.cell(row=2, column=2).value == result["summary"]["total_revenue"]


def test_build_arg_parser_requires_input():
    parser = build_arg_parser()

    args = parser.parse_args(["--input", "data.csv"])

    assert args.input == "data.csv"
    assert args.output_dir == "output"
