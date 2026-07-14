import openpyxl
import pytest

from report_automation.exceptions import ReportGenerationError
from report_automation.report_generator import generate_excel_report, generate_html_report

SUMMARY = {
    "total_revenue": 2768.75,
    "total_transactions": 4,
    "average_ticket": 692.19,
    "unique_products": 3,
}
CATEGORIES = [
    {"category": "Electrónica", "revenue": 2705.0, "count": 3},
    {"category": "Oficina", "revenue": 63.75, "count": 1},
]
PRODUCTS = [
    {"product": "Laptop", "revenue": 2550.0},
    {"product": "Mouse", "revenue": 155.0},
]


def test_generate_excel_report_writes_expected_sheets(tmp_path):
    output_path = tmp_path / "reportes" / "reporte.xlsx"

    result_path = generate_excel_report(SUMMARY, CATEGORIES, PRODUCTS, output_path)

    assert result_path.exists()
    workbook = openpyxl.load_workbook(result_path)
    assert workbook.sheetnames == ["Resumen", "Categorías", "Top Productos"]

    category_sheet = workbook["Categorías"]
    assert category_sheet.cell(row=2, column=1).value == "Electrónica"
    assert category_sheet.cell(row=2, column=2).value == 2705.0

    product_sheet = workbook["Top Productos"]
    assert product_sheet.cell(row=2, column=1).value == "Laptop"


def test_generate_excel_report_invalid_path_raises(tmp_path):
    blocker = tmp_path / "blocker"
    blocker.write_text("soy un archivo, no una carpeta", encoding="utf-8")
    output_path = blocker / "reporte.xlsx"

    with pytest.raises(ReportGenerationError):
        generate_excel_report(SUMMARY, CATEGORIES, PRODUCTS, output_path)


def test_generate_html_report_contains_summary_data(tmp_path):
    output_path = tmp_path / "reportes" / "reporte.html"

    result_path = generate_html_report(SUMMARY, CATEGORIES, PRODUCTS, output_path)

    content = result_path.read_text(encoding="utf-8")
    assert "Reporte de ventas" in content
    assert "Electrónica" in content
    assert "Laptop" in content
    assert "2768.75" in content


def test_generate_html_report_invalid_path_raises(tmp_path):
    blocker = tmp_path / "blocker"
    blocker.write_text("soy un archivo, no una carpeta", encoding="utf-8")
    output_path = blocker / "reporte.html"

    with pytest.raises(ReportGenerationError):
        generate_html_report(SUMMARY, CATEGORIES, PRODUCTS, output_path)
