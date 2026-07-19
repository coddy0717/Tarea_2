"""Pruebas de regresión: fijan los resultados esperados para el dataset de
referencia (data/sample_transactions.csv). Si alguna de estas pruebas falla
tras un cambio de código, hay que verificar que el cambio en los cálculos
o en el contenido de los reportes sea intencional antes de actualizar los
valores esperados.
"""

from pathlib import Path

import openpyxl
import pytest

from report_automation.data_loader import load_transactions
from report_automation.processor import (
    calculate_summary,
    group_by_category,
    monthly_trend,
    top_products,
)
from report_automation.report_generator import generate_excel_report, generate_html_report

REFERENCE_DATASET = Path(__file__).resolve().parent.parent / "data" / "sample_transactions.csv"


@pytest.fixture(scope="module")
def reference_transactions():
    transactions, errors = load_transactions(REFERENCE_DATASET)
    assert errors == [], "El dataset de referencia no debería tener filas inválidas"
    return transactions


def test_reference_dataset_summary_is_stable(reference_transactions):
    summary = calculate_summary(reference_transactions)

    assert summary == {
        "total_revenue": 4085.0,
        "total_transactions": 10,
        "average_ticket": 408.5,
        "unique_products": 5,
    }


def test_reference_dataset_category_breakdown_is_stable(reference_transactions):
    breakdown = group_by_category(reference_transactions)

    assert breakdown == [
        {"category": "Electrónica", "revenue": e3, "count": 4},
        {"category": "Oficina", "revenue": 748.75, "count": 4},
        {"category": "Hogar", "revenue": 243.75, "count": 2},
    ]


def test_reference_dataset_top_products_is_stable(reference_transactions):
    ranking = top_products(reference_transactions, limit=3)

    assert ranking == [
        {"product": "Laptop", "revenue": 2550.0},
        {"product": "Silla ergonómica", "revenue": 600.0},
        {"product": "Mouse", "revenue": 542.5},
    ]


def test_reference_dataset_monthly_trend_is_stable(reference_transactions):
    trend = monthly_trend(reference_transactions)

    assert trend == [
        {"month": "2025-01", "revenue": 2300.0},
        {"month": "2025-02", "revenue": 1063.75},
        {"month": "2025-03", "revenue": 721.25},
    ]


def test_reference_dataset_excel_report_content_is_stable(reference_transactions, tmp_path):
    summary = calculate_summary(reference_transactions)
    categories = group_by_category(reference_transactions)
    products = top_products(reference_transactions)

    output_path = generate_excel_report(
        summary, categories, products, tmp_path / "reporte_ventas.xlsx"
    )

    workbook = openpyxl.load_workbook(output_path)
    summary_sheet = workbook["Resumen"]
    assert summary_sheet.cell(row=2, column=2).value == 4085.0
    assert summary_sheet.cell(row=3, column=2).value == 10

    category_sheet = workbook["Categorías"]
    assert category_sheet.cell(row=2, column=1).value == "Electrónica"
    assert category_sheet.cell(row=2, column=2).value == 3092.5


def test_reference_dataset_html_report_content_is_stable(reference_transactions, tmp_path):
    summary = calculate_summary(reference_transactions)
    categories = group_by_category(reference_transactions)
    products = top_products(reference_transactions)

    output_path = generate_html_report(
        summary, categories, products, tmp_path / "reporte_ventas.html"
    )

    content = output_path.read_text(encoding="utf-8")
    assert "Ingreso total: 4085.0" in content
    assert "Electrónica" in content
    assert "Laptop" in content
