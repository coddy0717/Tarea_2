"""Pruebas end-to-end: ejecutan la aplicación como la usaría un usuario real
(subproceso de línea de comandos), sin importar las funciones internas.
"""

import os
import subprocess
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = PROJECT_ROOT / "src"

VALID_CSV = """date,category,product,quantity,unit_price
2025-01-05,Electrónica,Laptop,2,850.00
2025-01-08,Electrónica,Mouse,10,15.50
2025-01-09,Oficina,Papel,20,4.25
2025-02-02,Electrónica,Laptop,1,850.00
"""


def run_cli(args, cwd=None):
    # PYTHONPATH garantiza que "report_automation" se pueda importar sin
    # importar desde qué carpeta se ejecute el subproceso (p. ej. al
    # verificar la carpeta de salida por defecto desde un cwd distinto).
    env = os.environ.copy()
    env["PYTHONPATH"] = str(SRC_DIR)
    return subprocess.run(
        [sys.executable, "-m", "report_automation.main", *args],
        cwd=cwd or SRC_DIR,
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )


def test_cli_end_to_end_generates_valid_reports(tmp_path):
    csv_path = tmp_path / "transactions.csv"
    csv_path.write_text(VALID_CSV, encoding="utf-8")
    output_dir = tmp_path / "output"

    result = run_cli(["--input", str(csv_path), "--output-dir", str(output_dir)])

    assert result.returncode == 0, result.stderr
    assert "Reporte generado" in result.stdout

    excel_path = output_dir / "reporte_ventas.xlsx"
    html_path = output_dir / "reporte_ventas.html"
    assert excel_path.exists()
    assert html_path.exists()

    workbook = openpyxl.load_workbook(excel_path)
    assert workbook.sheetnames == ["Resumen", "Categorías", "Top Productos"]
    summary_sheet = workbook["Resumen"]
    assert summary_sheet.cell(row=2, column=1).value == "Ingreso total"
    assert summary_sheet.cell(row=2, column=2).value == 2790.0

    html_content = html_path.read_text(encoding="utf-8")
    assert "Reporte de ventas" in html_content
    assert "Electrónica" in html_content
    assert "Laptop" in html_content


def test_cli_end_to_end_uses_default_output_dir(tmp_path):
    csv_path = tmp_path / "transactions.csv"
    csv_path.write_text(VALID_CSV, encoding="utf-8")

    result = run_cli(["--input", str(csv_path)], cwd=tmp_path)

    assert result.returncode == 0, result.stderr
    assert (tmp_path / "output" / "reporte_ventas.xlsx").exists()
    assert (tmp_path / "output" / "reporte_ventas.html").exists()


def test_cli_end_to_end_fails_gracefully_for_missing_input(tmp_path):
    missing_csv = tmp_path / "no_existe.csv"

    result = run_cli(["--input", str(missing_csv)])

    assert result.returncode == 1
    assert "[ERROR]" in result.stderr
