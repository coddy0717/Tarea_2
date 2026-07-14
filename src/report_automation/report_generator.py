"""Generación de reportes de salida (Excel y HTML) a partir de los KPIs.

Funcionalidad a cargo del Equipo B: presentación y exportación de los
indicadores calculados por el módulo `processor`.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .exceptions import ReportGenerationError

HEADER_FONT = Font(bold=True)

SUMMARY_LABELS = {
    "total_revenue": "Ingreso total",
    "total_transactions": "Número de transacciones",
    "average_ticket": "Ticket promedio",
    "unique_products": "Productos distintos",
}


def generate_excel_report(summary, category_breakdown, top_products, output_path):
    """Genera un reporte Excel (.xlsx) con hojas de resumen, categorías y top productos."""
    try:
        workbook = Workbook()
        _write_summary_sheet(workbook.active, summary)
        _write_category_sheet(workbook.create_sheet("Categorías"), category_breakdown)
        _write_top_products_sheet(workbook.create_sheet("Top Productos"), top_products)

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
    except OSError as exc:
        raise ReportGenerationError(f"No se pudo generar el reporte Excel: {exc}") from exc

    return path


def _write_summary_sheet(sheet, summary):
    sheet.title = "Resumen"
    sheet.append(["Indicador", "Valor"])
    for cell in sheet[1]:
        cell.font = HEADER_FONT

    for key, label in SUMMARY_LABELS.items():
        sheet.append([label, summary.get(key)])


def _write_category_sheet(sheet, category_breakdown):
    sheet.append(["Categoría", "Ingreso", "Transacciones"])
    for cell in sheet[1]:
        cell.font = HEADER_FONT
    for item in category_breakdown:
        sheet.append([item["category"], item["revenue"], item["count"]])


def _write_top_products_sheet(sheet, top_products):
    sheet.append(["Producto", "Ingreso"])
    for cell in sheet[1]:
        cell.font = HEADER_FONT
    for item in top_products:
        sheet.append([item["product"], item["revenue"]])


def generate_html_report(summary, category_breakdown, top_products, output_path):
    """Genera un reporte HTML simple y autocontenido con los KPIs calculados."""
    rows_categories = "".join(
        f"<tr><td>{item['category']}</td><td>{item['revenue']}</td><td>{item['count']}</td></tr>"
        for item in category_breakdown
    )
    rows_products = "".join(
        f"<tr><td>{item['product']}</td><td>{item['revenue']}</td></tr>"
        for item in top_products
    )

    html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Reporte de ventas</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 2rem; }}
table {{ border-collapse: collapse; width: 100%; margin-bottom: 2rem; }}
th, td {{ border: 1px solid #ccc; padding: 0.5rem; text-align: left; }}
th {{ background-color: #f2f2f2; }}
</style>
</head>
<body>
<h1>Reporte de ventas</h1>
<h2>Resumen</h2>
<ul>
<li>Ingreso total: {summary.get('total_revenue')}</li>
<li>Transacciones: {summary.get('total_transactions')}</li>
<li>Ticket promedio: {summary.get('average_ticket')}</li>
<li>Productos distintos: {summary.get('unique_products')}</li>
</ul>
<h2>Ingresos por categoría</h2>
<table>
<tr><th>Categoría</th><th>Ingreso</th><th>Transacciones</th></tr>
{rows_categories}
</table>
<h2>Top productos</h2>
<table>
<tr><th>Producto</th><th>Ingreso</th></tr>
{rows_products}
</table>
</body>
</html>
"""

    try:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(html, encoding="utf-8")
    except OSError as exc:
        raise ReportGenerationError(f"No se pudo generar el reporte HTML: {exc}") from exc

    return path
